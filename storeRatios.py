
from pprint import pprint
import numpy
import pandas as pd
import csv
import json 
import requests
from datetime import datetime
from collections import Counter
from datetime import date
from my_dictionary import my_dictionary
from notify_run import Notify
import os
import xlrd
import xlsxwriter
import utils
from openpyxl import load_workbook


def getKey(ratio):
	switcher={
		'Face Value':4,
		'Dividend Per Share':5,
		'Operating Profit Per Share (Rs)':6,
		'Net Operating Profit Per Share (Rs)':7,
		'Free Reserves Per Share (Rs)':8,
		'Bonus in Equity Capital':9,
		'Operating Profit Margin(%)':10,
		'Profit Before Interest And Tax Margin(%)':11,
		'Gross Profit Margin(%)':12,
		'Cash Profit Margin(%)':13,
		'Adjusted Cash Margin(%)':14,
		'Net Profit Margin(%)':15,
		'Net Profit Margin':15,
		'Adjusted Net Profit Margin(%)':16,
		'Return On Capital Employed(%)':17,
		'Return On Net Worth(%)':18,
		'Return on Net Worth(%)':18,
		'Adjusted Return on Net Worth(%)':19,
		'Return on Assets Excluding Revaluations':20,
		'Return on Assets Including Revaluations':21,
		'Return on Long Term Funds(%)':22,
		'Return on Long Term Fund(%)':22,
		'Current Ratio':23,
		'Quick Ratio':24,
		'Debt Equity Ratio':25,
		'Long Term Debt Equity Ratio':26,
		'Interest Cover':27,
		'Total Debt to Owners Fund':28,
		'Financial Charges Coverage Ratio':29,
		'Financial Charges Coverage Ratio Post Tax':30,
		'Inventory Turnover Ratio':31,
		'Debtors Turnover Ratio':32,
		'Loans Turnover':32,
		'Investments Turnover Ratio':33,
		'Fixed Assets Turnover Ratio':34,
		'Total Assets Turnover Ratio':35,
		'Total Assets Turnover Ratios':35,
		'Asset Turnover Ratio':36,
		'Average Raw Material Holding':37,
		'Average Finished Goods Held':38,
		'Number of Days In Working Capital':39,
		'Material Cost Composition':40,
		'Imported Composition of Raw Materials Consumed':41,
		'Selling Distribution Cost Composition':42,
		'Expenses as Composition of Total Sales':43,
		'Dividend Payout Ratio Net Profit':44,
		'Dividend Payout Ratio Cash Profit':45,
		'Earning Retention Ratio':46,
		'Cash Earning Retention Ratio':47,
		'AdjustedCash Flow Times':48,
		'Earnings Per Share':49,
		'Book Value':50,
		'Cash Deposit Ratio':51,
		'Investment Deposit Ratio':52,
		'Credit Deposit Ratio':53,
		'Advances / Loans Funds(%)':54,
		'Capital Adequacy Ratio':55,
		'Operating Expense / Total Income':56,
		'Other Income / Total Income':57,
		'Interest Expended / Interest Earned':58,
		'Interest Expended / Capital Employed(%)':59,
		'Total Income / Capital Employed(%)':60,
		'Interest Spread': 61,
		'Return on Long Term Fund(%)':62,
		'Interest Income / Total Funds':63,
		'Net Interest Income / Total Funds':64,
		'Non Interest Income / Total Funds':65,
		'Interest Expended / Total Funds':66,
		'Operating Expense / Total Funds':67,
		'Profit Before Provisions / Total Funds':68,
		'Net Profit / Total Funds':69
	}
	return switcher.get(ratio, -1)


def main():
	#main function
	
	
	#read list of all stock
	df = utils.readExcel('stock-unique.xlsx')
		
	headers = {'authorization': "Basic API Key Ommitted", 'accept': "application/json", 'accept': "text/csv"}

	dataDf = pd.DataFrame(index=['id'], columns=['currentPrice', 'ratioData', 'industry'])

	wb = load_workbook("Ratios.xlsx")
	wbHeaders = ['Share', 'Industry', 'Year', 'Month','Face Value','Dividend Per Share','Operating Profit Per Share (Rs)','Net Operating Profit Per Share (Rs)','Free Reserves Per Share (Rs)', 'Bonus in Equity Capital','Operating Profit Margin(%)','Profit Before Interest And Tax Margin(%)',	'Gross Profit Margin(%)','Cash Profit Margin(%)','Adjusted Cash Margin(%)','Net Profit Margin(%)','Adjusted Net Profit Margin(%)','Return On Capital Employed(%)','Return On Net Worth(%)','Adjusted Return on Net Worth(%)','Return on Assets Excluding Revaluations','Return on Assets Including Revaluations','Return on Long Term Funds(%)','Current Ratio','Quick Ratio','Debt Equity Ratio','Long Term Debt Equity Ratio','Interest Cover','Total Debt to Owners Fund','Financial Charges Coverage Ratio','Financial Charges Coverage Ratio Post Tax','Inventory Turnover Ratio',	'Debtors Turnover Ratio','Investments Turnover Ratio', 'Fixed Assets Turnover Ratio','Total Assets Turnover Ratio','Asset Turnover Ratio', 'Average Raw Material Holding', 'Average Finished Goods Held', 'Number of Days In Working Capital','Material Cost Composition','Imported Composition of Raw Materials Consumed',	'Selling Distribution Cost Composition','Expenses as Composition of Total Sales', 'Dividend Payout Ratio Net Profit', 'Dividend Payout Ratio Cash Profit', 'Earning Retention Ratio', 'Cash Earning Retention Ratio', 'AdjustedCash Flow Times', 'Earnings Per Share', 'Book Value','Cash Deposit Ratio', 'Investment Deposit Ratio', 'Credit Deposit Ratio', 'Advances / Loans Funds(%)', 'Capital Adequacy Ratio', 'Operating Expense / Total Income', 'Other Income / Total Income', 'Interest Expended / Interest Earned', 'Interest Expended / Capital Employed(%)', 'Total Income / Capital Employed(%)', 'Interest Spread','Interest Income / Total Funds','Net Interest Income / Total Funds','Non Interest Income / Total Funds','Interest Expended / Total Funds', 'Operating Expense / Total Funds', 'Profit Before Provisions / Total Funds','Net Profit / Total Funds']
	
	
	# Select First Worksheet
	#ws = wb.worksheets[0]
	wb.remove(wb.worksheets[0])
	wb.create_sheet('Ratios', 0)
	ws = wb.worksheets[0]

	ws.append(wbHeaders)
	
	#Loop through all, add to topBuyList
	for index, row in df.iterrows():
		url = 'https://appfeeds.moneycontrol.com//jsonapi//stocks//graph&format=json&range=max&type=area&ex=&sc_id='+str(row['id'])
		rcomp = requests.get(url, headers=headers)
		data = json.loads(rcomp.text)
		currentPrice = float(data['graph']['current_close'])
		
		ratioUrl = 'https://appfeeds.moneycontrol.com//jsonapi//stocks//ratios&type=standalone&scid='+str(row['id'])
		ratioComp = requests.get(ratioUrl, headers=headers)
		rData = json.loads(ratioComp.text)
		
		for item in rData['company_data']['ratios']:
		
			row_data = [None] * 70
			row_data[0] = str(row['id'])
			row_data[1] = str(row['Industry'])
			row_data[2] = str(item['year'])
			row_data[3] = str(item['month'])
			
			for ratio in item ['item']:
				index = getKey(ratio['name'])
				
				if index == -1:
					print str(ratio['name'])+':'+str(ratio['value'])
				else:
					row_data[index] = str(ratio['value'])

		
			# Append Row Values
			ws.append(row_data)

		wb.save("Ratios.xlsx")

	
if __name__ == "__main__":
    main()
	