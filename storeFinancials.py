
from pprint import pprint
import numpy
import pandas as pd
import csv
import json 
import requests
from datetime import datetime
from collections import Counter
from datetime import date
from my_dictionary import my_dictionary
from notify_run import Notify
import os
import xlrd
import xlsxwriter
import utils
from openpyxl import load_workbook


def getKey(ratio):
	switcher={
		'Non-encumbered':4,
		'Net Sales/Income from operations':5,
		'Other Operating Income':6,
		'Total Income From Operations':7,
		'EXPENDITURE':8,
		'Consumption of Raw Materials':9,
		'Purchase of Traded Goods':10,
		'Increase/Decrease in Stocks':11,
		'Power & Fuel':12,
		'Employees Cost':13,
		'Depreciation':14,
		'Excise Duty ':15,
		'Admin. And Selling Expenses ':16,
		'R & D Expenses ':17,
		'Provisions And Contingencies ':18,
		'Provisions And Contingencies':18,
		'Exp. Capitalised':19,
		'Other Expenses':20,
		'P/L Before Other Inc. , Int., Excpt. Items & Tax':21,
		'Other Income':22,
		'P/L Before Int., Excpt. Items & Tax':23,
		'Interest':24,
		'P/L Before Exceptional Items & Tax':25,
		'Exceptional Items':26,
		'P/L Before Tax':27,
		'Tax':28,
		'P/L After Tax from Ordinary Activities':29,
		'P/L After Tax from Ordinary Activities ':29,
		'Prior Year Adjustments ':30,
		'Extra Ordinary Items':31,
		'Net Profit/(Loss) For the Period':32,
		'Equity Share Capital':33,
		'Reserves Excluding Revaluation Reserves':34,
		'Reserves Excluding Revaluation Reserves ':34,
		'Equity Dividend Rate (%)':35,
		'Basic EPS':36,
		'Diluted EPS':37,
		'EPS Before Extra Ordinary':38,
		'EPS After Extra Ordinary':39,
		'Public Share Holding':40,
		'- Number of shares (Crores)':41,
		'No Of Shares (Crores)':41,
		'Share Holding (%)':42,
		'Promoters and Promoter Group Shareholding':43,
		'a) Pledged/Encumbered':44,
		'Pledged/Encumbered':44,
		'- Per. of shares (as a % of the total sh. of prom. and promoter group)':45,
		'- Per. of shares (as a % of the total Share Cap. of the company)':46,
		'- Per. of shares (as a % of the total sh. of prom. and promoter group)':47,
		'- Per. of shares (as a % of the total Share Cap. of the company)':48,
		'Return on Assets %':49,
		'Operating Profit before Provisions and contingencies':50,
		'% of Share by Govt.':51,
		'Capital Adequacy Ratio - Basel - II':52,
		'Capital Adequacy Ratio - Basel -II':52,
		'Gross NPA':53,
		'Net NPA':54,
		'% of Gross NPA':55,
		'% of Net NPA':56,
		'Int. /Disc. on Adv/Bills':57,
		'Income on Investment':58,
		'Int. on balances With RBI':59,
		'Others':60,
		'Interest Expended':61
	}
	return switcher.get(ratio, -1)


def main():
	#main function
	print 'Storing financials'
	
	#read list of all stock
	df = utils.readExcel('stock-unique.xlsx')
		
	headers = {'authorization': "Basic API Key Ommitted", 'accept': "application/json", 'accept': "text/csv"}

	wb = load_workbook("Financials.xlsx")
	wbHeaders = ['Share', 'Industry', 'Year', 'Month', 'Non-encumbered', 'Net Sales/Income from operations', 'Other Operating Income', 'Total Income From Operations', 'EXPENDITURE', 'Consumption of Raw Materials', 'Purchase of Traded Goods', 'Increase/Decrease in Stocks', 'Power & Fuel', 'Employees Cost', 'Depreciation', 'Excise Duty ', 'Admin. And Selling Expenses ', 'R & D Expenses ', 'Operating Profit before Provisions and contingencies','Provisions And Contingencies ', 'Exp. Capitalised', 'Other Expenses', 'P/L Before Other Inc. , Int., Excpt. Items & Tax', 'Other Income', 'P/L Before Int., Excpt. Items & Tax', 'Interest', 'P/L Before Exceptional Items & Tax', 'Exceptional Items', 'P/L Before Tax', 'Tax', 'P/L After Tax from Ordinary Activities', 'Prior Year Adjustments ', 'Extra Ordinary Items', 'Net Profit/(Loss) For the Period', 'Equity Share Capital', 'Reserves Excluding Revaluation Reserves', 'Equity Dividend Rate (%)', 'Basic EPS', 'Diluted EPS', 'EPS Before Extra Ordinary', 'EPS After Extra Ordinary', 'Number of shares (Crores)', 'Share Holding (%)', 'Public Share Holding', 'Promoters and Promoter Group Shareholding',  'a) Pledged/Encumbered' 'Pledged/Encumbered', '- Per. of shares (as a % of the total sh. of prom. and promoter group)', '- Per. of shares (as a % of the total Share Cap. of the company)', '- Per. of shares (as a % of the total sh. of prom. and promoter group)','- Per. of shares (as a % of the total Share Cap. of the company)','% of Share by Govt.','Capital Adequacy Ratio - Basel - II','Gross NPA','Net NPA', '% of Gross NPA','% of Net NPA', 'Return on Assets %', 'Int. /Disc. on Adv/Bills', 'Income on Investment', 'Int. on balances With RBI', 'Others', 'Interest Expended']
	
	
	# Select First Worksheet
	#ws = wb.worksheets[0]
	wb.remove(wb.worksheets[0])
	wb.create_sheet('Ratios', 0)
	ws = wb.worksheets[0]

	ws.append(wbHeaders)
	
	#Loop through all, add to topBuyList
	for index, row in df.iterrows():
		try:
			#print 'storing for '+str(row['id'])
			ratioUrl = 'https://appfeeds.moneycontrol.com/jsonapi/stocks/finacials&format=&section=quaterly&type=standalone&scid='+str(row['id'])
			ratioComp = requests.get(ratioUrl, headers=headers)
			rData = json.loads(ratioComp.text)
			print 'storing for '+str(row['id'])
			for item in rData['company_data']['result']:
			
				row_data = [None] * 62
				row_data[0] = str(row['id'])
				row_data[1] = str(row['Industry'])
				row_data[2] = str(item['year'])
				row_data[3] = str(item['month'])
				
				for result in item ['item']:
					index = getKey(result['name'])
					
					if index == -1 and int(result['head_flag']) is not 1:
						print str(result['name'])+':'+str(result['value'])
					else:
						row_data[index] = (result['value']).replace(',', '')

			
				# Append Row Values
				ws.append(row_data)
		except Exception as e:
			print e

		wb.save("Financials.xlsx")

	
if __name__ == "__main__":
    main()
	