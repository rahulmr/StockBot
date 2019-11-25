
from pprint import pprint
import numpy as np
import pandas as pd
import csv
import json 
import requests
from datetime import datetime
from collections import Counter
from datetime import date
from my_dictionary import my_dictionary
from notify_run import Notify
import os
import xlrd
import utils
from openpyxl import load_workbook

stockList = []

today = date.today()

#function to return list of stock names
def readExcel():
	df = pd.read_excel('stock-unique.xlsx', sheet_name=0, keep_default_na=False) # can also index sheet by name or fetch all sheets
	stockList = df['id'].tolist() 
	return stockList
	
#function to compare averages of the stock, add to topBuyList if high, otherwise check if it is part of bought list
def storeNews(stockList):

	
	wb = load_workbook("News.xlsx")
	wbHeaders = ['Stock', 'Date', 'News', 'Message']
	wb.remove(wb.worksheets[0])
	wb.create_sheet('News', 0)
	ws = wb.worksheets[0]
	ws.append(wbHeaders)
	
	#Loop through all, add to topBuyList
	for item in stockList:
		print 'Stock is '+item
		url = 'https://www.moneycontrol.com/news18/stocks/overview/'+str(item)+'/N'
		headers = {'authorization': "Basic API Key Ommitted", 'accept': "application/json", 'accept': "text/csv"}

		rcomp = requests.get(url, headers=headers)
		data = json.loads(rcomp.text)
	
		try:
			items = data['alerts']
			for row in items: 
				row_data = [None] * 4
				row_data[0] = str(item)
				row_data[1] = str(row.get('title'))
				row_data[2] = str(row.get('entdate'))
				row_data[3] = str(row.get('message'))
				ws.append(row_data)
			
		except Exception as e:
			print e
	
	wb.save("News.xlsx")


def main():
	#main function
	
	#read list of all stock
	stockList = readExcel()
		
		
	#Loop through all, add to topBuyList
	storeNews(stockList)
	
	
if __name__ == "__main__":
    main()
	